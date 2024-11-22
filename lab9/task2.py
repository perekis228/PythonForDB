import openpyxl

workbook = openpyxl.load_workbook('lab9.xlsx')
sheet = workbook.active

max_pay = 0.0
max_pay_worker = None
max_pay_dept = None

min_pay = 10.0**10
min_pay_worker = None
min_pay_dept = None

mid_salaries = dict()

men = 0
men_dept = None
for row in range(2, 13):
    if sheet.cell(row=row, column=1).value is not None:
        if sheet.cell(row=row, column=9).value > max_pay:
            max_pay = sheet.cell(row=row, column=9).value
            max_pay_worker = sheet.cell(row=row, column=2).value
            max_pay_dept = sheet.cell(row=row, column=3).value
        elif sheet.cell(row=row, column=9).value < min_pay:
            min_pay = sheet.cell(row=row, column=9).value
            min_pay_worker = sheet.cell(row=row, column=2).value
            min_pay_dept = sheet.cell(row=row, column=3).value

        men += 1
        men_dept = sheet.cell(row=row, column=3).value
    else:
        if men != 0:
            mid_salaries[men_dept] = sheet.cell(row=row, column=9).value / men
        men = 0

sheet.cell(row=14, column=2, value='Мин. зарплата:')
sheet.cell(row=14, column=3, value=min_pay_worker)
sheet.cell(row=14, column=4, value=min_pay_dept)
sheet.cell(row=14, column=5, value=min_pay).number_format = '#,##0.00"₽"'

sheet.cell(row=15, column=2, value='Макс. зарплата:')
sheet.cell(row=15, column=3, value=max_pay_worker)
sheet.cell(row=15, column=4, value=max_pay_dept)
sheet.cell(row=15, column=5, value=max_pay).number_format = '#,##0.00"₽"'

sheet.cell(row=16, column=2, value='Средняя зарплата:')
for i, dept in enumerate(mid_salaries.keys()):
    sheet.cell(row=16+i, column=3, value=dept)
    sheet.cell(row=16+i, column=4, value=mid_salaries[dept]).number_format = '#,##0.00"₽"'

#Красота:
for column in sheet.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column[1:]:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)
    sheet.column_dimensions[column_letter].width = adjusted_width

workbook.save('lab9.xlsx')
workbook.close()