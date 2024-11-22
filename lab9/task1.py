import openpyxl
from openpyxl.styles import Alignment

workers = [{'Номер': '0002', 'ФИО': 'Петров П.П.',  'Отдел': 'Бухгалтерия',  'Оклад': 3913.04, 'Надбавка': 2608.70, 'НДФЛ': 0.13},
           {'Номер': '0005', 'ФИО': 'Васин В.В.',   'Отдел': 'Бухгалтерия',  'Оклад': 5934.78, 'Надбавка': 913.04,  'НДФЛ': 0.13},
           {'Номер': '0001', 'ФИО': 'Иванов И.И.',  'Отдел': 'Отдел кадров', 'Оклад': 6000.00, 'Надбавка': 4000.00, 'НДФЛ': 0.13},
           {'Номер': '0003', 'ФИО': 'Сидоров С.С.', 'Отдел': 'Отдел кадров', 'Оклад': 5000.00, 'Надбавка': 4500.00, 'НДФЛ': 0.13},
           {'Номер': '0006', 'ФИО': 'Львов Л.Л.',   'Отдел': 'Отдел кадров', 'Оклад': 4074.07, 'Надбавка': 2444.44, 'НДФЛ': 0.13},
           {'Номер': '0007', 'ФИО': 'Волков В.В.',  'Отдел': 'Отдел кадров', 'Оклад': 1434.78, 'Надбавка': 1434.78, 'НДФЛ': 0.13},
           {'Номер': '0004', 'ФИО': 'Машин М.М.',   'Отдел': 'Столовая',     'Оклад': 5500.00, 'Надбавка': 3500.00, 'НДФЛ': 0.13}]
workbook = openpyxl.Workbook()
sheet = workbook.active

sheet['A1'] = 'Таб. номер'
sheet['B1'] = 'Фамилия'
sheet['C1'] = 'Отдел'
sheet['D1'] = 'Сумма по окладу, руб.'
sheet['E1'] = 'Сумма по надбавкам, руб.'
sheet['F1'] = 'Сумма зарплаты, руб.'
sheet['G1'] = 'НДФЛ, %'
sheet['H1'] = 'Сумма НДФЛ, руб.'
sheet['I1'] = 'Сумма к выдаче, руб.'

minus = 2
for row in range(2, 13):
    if row not in [4, 9, 11, 12]:
        sheet[f'A{row}'] = workers[row-minus]['Номер']
        sheet[f'B{row}'] = workers[row-minus]['ФИО']
        sheet[f'C{row}'] = workers[row-minus]['Отдел']
        sheet[f'D{row}'] = workers[row-minus]['Оклад']
        sheet[f'E{row}'] = workers[row-minus]['Надбавка']
        sheet[f'F{row}'] = workers[row-minus]['Оклад'] + workers[row-minus]['Надбавка']
        sheet[f'G{row}'] = workers[row-minus]['НДФЛ']
        sheet[f'H{row}'] = round((workers[row-minus]['Оклад'] + workers[row-minus]['Надбавка']) * workers[row-minus]['НДФЛ'], ndigits=2)
        sheet[f'I{row}'] = round((workers[row-minus]['Оклад'] + workers[row-minus]['Надбавка']) * (1-workers[row-minus]['НДФЛ']), ndigits=2)
    elif row == 4:
        sheet[f'C4'] = 'Бухгалтерия\nИтог'
        sheet['C4'].alignment = Alignment(wrap_text=True)
        for col in range(4, 10):
            if col != 7:
                val = sum(sheet.cell(row=row, column=col).value for row in range(2, 4))
                sheet.cell(row=row, column=col, value=val)
        minus += 1
    elif row == 9:
        sheet[f'C9'] = 'Отдел кадров\nИтог'
        sheet['C9'].alignment = Alignment(wrap_text=True)
        for col in range(4, 10):
            if col != 7:
                val = sum(sheet.cell(row=row, column=col).value for row in range(5, 9))
                sheet.cell(row=row, column=col, value=val)
        minus += 1
    elif row == 11:
        sheet[f'C11'] = 'Столовая\nИтог'
        sheet['C11'].alignment = Alignment(wrap_text=True)
        for col in range(4, 10):
            if col != 7:
                val = sheet.cell(row=10, column=col).value
                sheet.cell(row=row, column=col, value=val)
    elif row == 12:
        sheet[f'C12'] = 'Общий итог'
        for col in range(4, 10):
            if col != 7:
                val = sum(sheet.cell(row=row, column=col).value for row in [4, 9, 11])
                sheet.cell(row=row, column=col, value=val)

#Красота:
for col in range(1, 10):
    sheet.cell(row=1, column=col).alignment = Alignment(vertical='center', text_rotation=90)

for row in range(2, 12):
    for col in range(4, 10):
        if col != 7:
            sheet.cell(row=row, column=col).number_format = '#,##0.00"₽"'
        else:
            sheet.cell(row=row, column=col).number_format = '0%'

for column in sheet.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column[1:]:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)
    sheet.column_dimensions[column_letter].width = adjusted_width

workbook.save('lab9.xlsx')
workbook.close()